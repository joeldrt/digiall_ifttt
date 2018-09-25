from flask import request, send_from_directory
from flask_restful import Resource, reqparse
from flask_jwt_extended import (jwt_required, get_jwt_identity)
from openpyxl import Workbook

from data.registro_sensor import RegistroSensor
from services import registro_sensor_service, habitacion_service


obtener_registros_por_habitacion_parser = reqparse.RequestParser(bundle_errors=True)
obtener_registros_por_habitacion_parser.add_argument('fecha_inicial', type=str, location='args')
obtener_registros_por_habitacion_parser.add_argument('fecha_final', type=str, location='args')


class ObtenerRegistrosPorHabitacion(Resource):
    @jwt_required
    def get(self, habitacion_id):
        usuario_propietario = get_jwt_identity()

        if not habitacion_service.habitacion_le_pertenece_a_propietario(usuario_propietario=usuario_propietario,
                                                                        habitacion_id=habitacion_id):
            return {'message': 'La habitación no es del propietario'}, 403

        data = obtener_registros_por_habitacion_parser.parse_args()

        fecha_inicial = data['fecha_inicial']
        fecha_final = data['fecha_final']

        try:
            registros_objs = registro_sensor_service.obtener_todos_entre_fechas_por_habitacion(habitacion_id=habitacion_id,
                                                                                              fecha_inicial=fecha_inicial,
                                                                                              fecha_final=fecha_final)

            registros = [
                registro.to_dict() for registro in registros_objs
            ]
        except Exception as exception:
            return {'message': 'Error del servidor al obtener los registros por habitación'}, 500

        return registros


class CreateExcelFile(Resource):
    def get(self):
        fecha_inicial = str(request.args.get('fecha_inicial'))
        fecha_final = str(request.args.get('fecha_final'))
        complejo = str(request.args.get('complejo'))
        habitacion = str(request.args.get('habitacion'))

        registros = registro_sensor_service.obtener_todos_entre_fechas_por_habitacion()

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='id registro')
        ws.cell(row=1, column=2, value='fecha')
        ws.cell(row=1, column=3, value='tipo de evento')
        ws.cell(row=1, column=4, value='complejo')
        ws.cell(row=1, column=5, value='habitacion')
        ws.cell(row=1, column=6, value='sensor')

        for row, registro in enumerate(registros):
            ws.cell(row=(row + 2), column=1, value=registro['id'])
            ws.cell(row=(row + 2), column=2, value=registro['fecha_creacion'])
            ws.cell(row=(row + 2), column=3, value=registro['tipo_evento'])
            ws.cell(row=(row + 2), column=4, value=registro['complejo'])
            ws.cell(row=(row + 2), column=5, value=registro['habitacion'])
            ws.cell(row=(row + 2), column=6, value=registro['sensor'])

        file_name = '{}_{}_{}_{}_{}.xlsx'.format(fecha_inicial, fecha_final, complejo, habitacion)

        wb.save('./static/{}'.format(file_name))

        return send_from_directory('static', file_name)
