from flask import request, send_from_directory
from flask_restful import Resource

from datetime import datetime, timedelta

from data.registro_sensor import RegistroSensor
from mongoengine.queryset.visitor import Q

from openpyxl import Workbook


def get_registros(fecha_inicial: str, fecha_final: str, sensor_id: str) -> []:
    fecha_inicial_date = datetime.strptime(fecha_inicial, '%Y-%m-%d')
    fecha_final_date = datetime.strptime(fecha_final, '%Y-%m-%d') + timedelta(days=1)

    if sensor_id == 'TODOS':
        objetos = RegistroSensor.objects(
            Q(fecha_creacion__gte=fecha_inicial_date) &
            Q(fecha_creacion__lte=fecha_final_date)
        )
    else:
        objetos = RegistroSensor.objects(
            Q(sensor_id=sensor_id) &
            Q(fecha_creacion__gte=fecha_inicial_date) &
            Q(fecha_creacion__lte=fecha_final_date)
        )

    registros = [
        registro.to_dict() for registro in objetos
    ]
    return registros


class ObtenerSensoresIds(Resource):
    def get(self):
        sensores_ids = RegistroSensor.objects().distinct('sensor_id')
        return sensores_ids


class MostrarRegistros(Resource):
    def get(self):
        fecha_inicial = str(request.args.get('fecha_inicial'))
        fecha_final = str(request.args.get('fecha_final'))
        sensor_id = str(request.args.get('sensor_id'))

        registros = get_registros(fecha_inicial, fecha_final, sensor_id)

        return registros


class CreateExcelFile(Resource):
    def get(self):
        fecha_inicial = str(request.args.get('fecha_inicial'))
        fecha_final = str(request.args.get('fecha_final'))
        sensor_id = str(request.args.get('sensor_id'))

        registros = get_registros(fecha_inicial, fecha_final, sensor_id)

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='id registro')
        ws.cell(row=1, column=2, value='fecha')
        ws.cell(row=1, column=3, value='tipo de evento')
        ws.cell(row=1, column=4, value='id sensor')

        for row, registro in enumerate(registros):
            ws.cell(row=(row + 2), column=1, value=registro['id'])
            ws.cell(row=(row + 2), column=2, value=registro['fecha_creacion'])
            ws.cell(row=(row + 2), column=3, value=registro['tipo_evento'])
            ws.cell(row=(row + 2), column=4, value=registro['sensor_id'])

        file_name = '{}_{}_{}.xlsx'.format(fecha_inicial, fecha_final, sensor_id)

        wb.save('./static/{}'.format(file_name))

        return send_from_directory('static', file_name)
