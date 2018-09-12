from flask import Flask, request, Response
from flask_restful import Resource, Api, reqparse
from flask_cors import CORS
from datetime import datetime, timedelta
import logging
import json

import data.mongo_setup as mongo_setup

app_ifttt_channel_key = 'dV8XXrIUIoBe9jK04jlXA0hkzf1El4wBjbaW9VezxnCxZAINw7bnoFxfej-9fWBk'
ACTIVADO = 'SENSOR_ACTIVADO'
NORMAL = 'SENSOR_REGRESO_NORMALIDAD'

app = Flask(__name__, static_url_path='/static')
handler = logging.FileHandler('da_ifttt_api.log')
handler.setLevel(logging.DEBUG)
app.logger.addHandler(handler)

CORS(app)
api = Api(app)

mongo_setup.global_init()

# parser = reqparse.RequestParser()
# parser.add_argument('actionFields', type=dict, required=True)
# parser.add_argument('ifttt_source', type=dict)
# parser.add_argument('user', type=dict)

from data.registro_sensor import RegistroSensor
from flask import send_from_directory
from mongoengine.queryset.visitor import Q
from openpyxl import Workbook


class HelloWorld(Resource):
    def get(self):
        return {'hello': 'world'}


class SensorActivado(Resource):
    def post(self):
        
        def myconverter(o):
            if isinstance(o, datetime):
                return o.__str__()

        errors_array = list()

        channel_key = request.headers['IFTTT-Channel-Key']
        if channel_key != app_ifttt_channel_key:
            error = {'message': 'INVALID IFTTT-Channel-Key'}
            errors_array.append(error)
            error_object = {'errors': errors_array}

            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=401)

            return resp
        
        data = request.get_json()
        
        print("{} SENSOR ACTIVADO ----------------".format(datetime.now()))

        if 'actionFields' not in dict(data).keys():
            error = {'message': 'actionFields key is missing'}
            errors_array.append(error)
            error_object = {'errors': errors_array}
            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=400)
            return resp
        
        if 'sensor_id' not in dict(data['actionFields']).keys():
            error = {'message': 'missing sensor_id value'}
            errors_array.append(error)
            error_object = {'errors': errors_array}
            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=400)
            return resp
        
        sensor_id = data['actionFields']['sensor_id']
        print("sensor_id: {}".format(sensor_id))
        
        if 'SENSOR-CUARTO-SKIP' == sensor_id:
            error = {'message': 'Sensor Id not registered'}
            errors_array.append(error)
            error2 = {'status': 'SKIP', 'message': 'ID_not accepted'}
            errors_array.append(error2)
            error_object = {'errors': errors_array}
            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=400)
            return resp
        
        if 'reported_at' not in dict(data['actionFields']).keys():
            error = {'message': 'missing reported_at value'}
            errors_array.append(error)
            error_object = {'errors': errors_array}
            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=400)
            return resp
        
        reported_at = data['actionFields']['reported_at']
        
        print("reported_at: {}".format(reported_at))

        registro_sensor = RegistroSensor()
        registro_sensor.fecha_creacion = datetime.now()
        registro_sensor.tipo_evento = ACTIVADO
        registro_sensor.sensor_id = sensor_id
        registro_sensor.reported_at = reported_at

        registro_sensor.save()

        response = {'id': str(registro_sensor.id),
                    'url': 'http://www.digiall.mx'}

        response_data = list()
        response_data.append(response)

        last_object = {'data': response_data}

        resp = Response(json.dumps(last_object, ensure_ascii=False, default=myconverter), mimetype='application/json; charset=utf-8')

        return resp


class SensorNormal(Resource):
    def post(self):

        def myconverter(o):
            if isinstance(o, datetime):
                return o.__str__()

        errors_array = list()

        channel_key = request.headers['IFTTT-Channel-Key']
        if channel_key != app_ifttt_channel_key:
            error = {'message': 'INVALID IFTTT-Channel-Key'}
            errors_array.append(error)
            error_object = {'errors': errors_array}

            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=401)

            return resp
        
        data = request.get_json()
        
        print("{} SENSOR REGRESANDO A LA NORMALIDAD ----------------".format(datetime.now()))

        if 'actionFields' not in dict(data).keys():
            error = {'message': 'actionFields key is missing'}
            errors_array.append(error)
            error_object = {'errors': errors_array}
            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=400)
            return resp
        
        if 'sensor_id' not in dict(data['actionFields']).keys():
            error = {'message': 'missing sensor_id value'}
            errors_array.append(error)
            error_object = {'errors': errors_array}
            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=400)
            return resp
        
        sensor_id = data['actionFields']['sensor_id']
        print("sensor_id: {}".format(sensor_id))
        
        if 'SENSOR-CUARTO-SKIP' == sensor_id:
            error = {'message': 'Sensor Id not registered'}
            errors_array.append(error)
            error2 = {'status': 'SKIP', 'message': 'ID_not accepted'}
            errors_array.append(error2)
            error_object = {'errors': errors_array}
            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=400)
            return resp
        
        if 'reported_at' not in dict(data['actionFields']).keys():
            error = {'message': 'missing reported_at value'}
            errors_array.append(error)
            error_object = {'errors': errors_array}
            resp = Response(json.dumps(error_object, ensure_ascii=False, default=myconverter),
                            mimetype='application/json; charset=utf-8', status=400)
            return resp
        
        reported_at = data['actionFields']['reported_at']
        
        print("reported_at: {}".format(reported_at))

        registro_sensor = RegistroSensor()
        registro_sensor.fecha_creacion = datetime.now()
        registro_sensor.tipo_evento = NORMAL
        registro_sensor.sensor_id = sensor_id
        registro_sensor.reported_at = reported_at

        registro_sensor.save()

        response = {'id': str(registro_sensor.id),
                    'url': 'http://www.digiall.mx'}

        response_data = list()
        response_data.append(response)

        last_object = {'data': response_data}

        resp = Response(json.dumps(last_object, ensure_ascii=False, default=myconverter), mimetype='application/json; charset=utf-8')

        return resp


class StatusEndPoint(Resource):
    def get(self):
        for key, value in request.headers:
            logging.debug('Headers-------------------')
            logging.debug('{}: {}'.format(key, value))

        errors_array = list()

        if not request.headers['IFTTT-Channel-Key']:
            error = {'message': 'no IFTTT-Channel-Key header'}
            errors_array.append(error)
            return {'errors': errors_array}, 401

        channel_key = request.headers['IFTTT-Channel-Key']
        if channel_key != app_ifttt_channel_key:
            error = {'message': 'INVALID IFTTT-Channel-Key'}
            errors_array.append(error)
            return {'errors': errors_array}, 401

        logging.debug('IFTTT-Channel-Key: {}'.format(channel_key))

        return {'data': {
                            'test_pass': True
                        }
                }


class TestSetupEndPoint(Resource):
    def post(self):
        for key, value in request.headers:
            logging.debug('Headers-------------------')
            logging.debug('{}: {}'.format(key, value))

        errors_array = list()

        if not request.headers['IFTTT-Channel-Key']:
            error = {'message': 'no IFTTT-Channel-Key header'}
            errors_array.append(error)
            return {'errors': errors_array}, 401

        channel_key = request.headers['IFTTT-Channel-Key']
        if channel_key != app_ifttt_channel_key:
            error = {'message': 'INVALID IFTTT-Channel-Key'}
            errors_array.append(error)
            return {'errors': errors_array}, 401

        logging.debug('IFTTT-Channel-Key: {}'.format(channel_key))

        response_data = {
                    "data": {
                        "samples": {
                            "actions": {
                                "sensoractivado": {
                                    "sensor_id": "SENSOR-CUARTO1",
                                    "reported_at": "EDIFICIO1"
                                    },
                                "sensornormal": {
                                    "sensor_id": "SENSOR-CUARTO1",
                                    "reported_at": "EDIFICIO1"
                                    }
                                },
                            "actionRecordSkipping": {
                                "testaction": {
                                    "sensor_id": "SENSOR-CUARTO-SKIP",
                                    "reported_at": "EDIFICIO-SKIP"
                                    }
                                }
                            }
                        }
                    }

        resp = Response(json.dumps(response_data, ensure_ascii=False), mimetype='application/json; charset=utf-8')

        return resp


class MostrarRegistros(Resource):
    def get(self):
        days = int(request.args.get('days'))

        date_to_retrieve = datetime.today() - timedelta(days=days)

        #  RegistroSensor.objects(Q(fecha_creacion__lte=date_to_retrieve))

        registros = [
                registro.to_dict() for registro in RegistroSensor.objects(Q(fecha_creacion__gte=date_to_retrieve))
            ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='id registro')
        ws.cell(row=1, column=2, value='fecha')
        ws.cell(row=1, column=3, value='tipo de evento')
        ws.cell(row=1, column=4, value='id sensor')

        for row, registro in enumerate(registros):
            ws.cell(row=(row + 2), column=1, value=registro['id'])
            ws.cell(row=(row + 2), column=2, value=registro['fecha_creacion'])
            ws.cell(row=(row + 2), column=3, value=registro['tipo_evento'])
            ws.cell(row=(row + 2), column=4, value=registro['sensor_id'])

        wb.save('./static/{}.xlsx'.format(str(date_to_retrieve)))

        return registros


class CreateExcelFile(Resource):
    def get(self):
        days = int(request.args.get('days'))

        date_to_retrieve = datetime.today() - timedelta(days=days)

        registros = [
            registro.to_dict() for registro in RegistroSensor.objects(Q(fecha_creacion__gte=date_to_retrieve))
        ]

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='id registro')
        ws.cell(row=1, column=2, value='fecha')
        ws.cell(row=1, column=3, value='tipo de evento')
        ws.cell(row=1, column=4, value='id sensor')

        for row, registro in enumerate(registros):
            ws.cell(row=(row + 2), column=1, value=registro['id'])
            ws.cell(row=(row + 2), column=2, value=registro['fecha_creacion'])
            ws.cell(row=(row + 2), column=3, value=registro['tipo_evento'])
            ws.cell(row=(row + 2), column=4, value=registro['sensor_id'])

        file_name = str(date_to_retrieve) + '.xlsx'

        wb.save('./static/{}'.format(file_name))

        return send_from_directory('static', file_name)


api.add_resource(HelloWorld, '/api/helloworld')
api.add_resource(SensorActivado, '/api/ifttt/v1/actions/sensoractivado')
api.add_resource(SensorNormal, '/api/ifttt/v1/actions/sensornormal')
api.add_resource(StatusEndPoint, '/api/ifttt/v1/status')
api.add_resource(TestSetupEndPoint, '/api/ifttt/v1/test/setup')

api.add_resource(MostrarRegistros, '/api/registros')
api.add_resource(CreateExcelFile, '/api/descargar_excel')
